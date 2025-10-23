"""
Excel Workbook Template Generator.

This module creates structured Excel workbooks for financial reports:
1. Summary sheet with key metrics
2. Income Statement sheet
3. Balance Sheet sheet
4. Cash Flow sheet
5. Ratios and Analysis sheet
6. Raw Data sheet
"""

from typing import List, Dict, Optional
from datetime import date
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

from src.models.schemas import FinancialMetric


class ExcelTemplateGenerator:
    """
    Generate Excel workbooks from financial data.
    
    Creates comprehensive workbooks with:
    - Summary dashboard
    - Financial statement sheets
    - Analysis and ratios
    - Raw data
    """
    
    def __init__(self):
        self.logger = logger.bind(component="excel_template_generator")
        self.wb = None
    
    def create_financial_workbook(
        self,
        company_name: str,
        report_period: date,
        metrics: List[FinancialMetric],
        derived_metrics: Optional[List[FinancialMetric]] = None,
        output_path: str = "financial_report.xlsx"
    ) -> str:
        """
        Create complete financial workbook.
        
        Args:
            company_name: Company name
            report_period: Reporting period end date
            metrics: List of financial metrics
            derived_metrics: List of derived metrics (ratios, etc.)
            output_path: Output file path
        
        Returns:
            Path to generated workbook
        """
        self.logger.info(
            "creating_financial_workbook",
            company=company_name,
            period=str(report_period)
        )
        
        # Create workbook
        self.wb = Workbook()
        
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Add sheets
        self._add_summary_sheet(company_name, report_period, metrics, derived_metrics)
        self._add_income_statement_sheet(metrics)
        self._add_balance_sheet_sheet(metrics)
        self._add_cash_flow_sheet(metrics)
        if derived_metrics:
            self._add_ratios_sheet(derived_metrics)
        self._add_raw_data_sheet(metrics, derived_metrics)
        
        # Save workbook
        self.wb.save(output_path)
        
        self.logger.info("financial_workbook_created", path=output_path)
        
        return output_path
    
    def _add_summary_sheet(
        self,
        company_name: str,
        report_period: date,
        metrics: List[FinancialMetric],
        derived_metrics: Optional[List[FinancialMetric]]
    ):
        """Add summary dashboard sheet."""
        ws = self.wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = company_name
        ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
        
        ws['A2'] = "Financial Summary Report"
        ws['A2'].font = Font(size=12, bold=True)
        
        ws['A3'] = f"Period Ending: {report_period.strftime('%B %d, %Y')}"
        ws['A3'].font = Font(italic=True)
        
        # Key metrics section
        row = 5
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        self._apply_header_style(ws[f'A{row}'])
        
        row += 1
        
        # Get latest period metrics
        latest_metrics = [m for m in metrics if m.period_end_date == report_period]
        
        # Revenue
        revenue = next((m for m in latest_metrics if "revenue" in m.metric_name.lower()), None)
        if revenue:
            ws[f'A{row}'] = "Revenue"
            ws[f'B{row}'] = float(revenue.value)
            ws[f'C{row}'] = f"{revenue.currency} ({revenue.scale})"
            row += 1
        
        # Net Income
        net_income = next((m for m in latest_metrics if "net_income" in m.metric_name.lower()), None)
        if net_income:
            ws[f'A{row}'] = "Net Income"
            ws[f'B{row}'] = float(net_income.value)
            ws[f'C{row}'] = f"{net_income.currency} ({net_income.scale})"
            row += 1
        
        # Total Assets
        total_assets = next((m for m in latest_metrics if "total_assets" in m.metric_name.lower()), None)
        if total_assets:
            ws[f'A{row}'] = "Total Assets"
            ws[f'B{row}'] = float(total_assets.value)
            ws[f'C{row}'] = f"{total_assets.currency} ({total_assets.scale})"
            row += 1
        
        # Total Equity
        total_equity = next((m for m in latest_metrics if "total_equity" in m.metric_name.lower()), None)
        if total_equity:
            ws[f'A{row}'] = "Total Equity"
            ws[f'B{row}'] = float(total_equity.value)
            ws[f'C{row}'] = f"{total_equity.currency} ({total_equity.scale})"
            row += 1
        
        # Key ratios (if available)
        if derived_metrics:
            row += 2
            ws[f'A{row}'] = "Key Ratios"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            self._apply_header_style(ws[f'A{row}'])
            row += 1
            
            latest_ratios = [m for m in derived_metrics if m.period_end_date == report_period]
            
            for ratio_metric in latest_ratios[:5]:  # Top 5 ratios
                ws[f'A{row}'] = ratio_metric.metric_name
                ws[f'B{row}'] = float(ratio_metric.value)
                row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
    
    def _add_income_statement_sheet(self, metrics: List[FinancialMetric]):
        """Add income statement sheet."""
        ws = self.wb.create_sheet("Income Statement")
        
        # Filter income statement metrics
        is_keywords = ["revenue", "sales", "cost", "gross_profit", "operating", "ebitda", "net_income", "earnings"]
        is_metrics = [
            m for m in metrics
            if any(kw in m.metric_name.lower() for kw in is_keywords)
        ]
        
        if not is_metrics:
            ws['A1'] = "No income statement data available"
            return
        
        # Create table
        self._create_statement_table(ws, "Income Statement", is_metrics)
    
    def _add_balance_sheet_sheet(self, metrics: List[FinancialMetric]):
        """Add balance sheet sheet."""
        ws = self.wb.create_sheet("Balance Sheet")
        
        # Filter balance sheet metrics
        bs_keywords = ["asset", "liability", "equity", "cash", "inventory", "receivable", "payable", "debt"]
        bs_metrics = [
            m for m in metrics
            if any(kw in m.metric_name.lower() for kw in bs_keywords)
        ]
        
        if not bs_metrics:
            ws['A1'] = "No balance sheet data available"
            return
        
        # Create table
        self._create_statement_table(ws, "Balance Sheet", bs_metrics)
    
    def _add_cash_flow_sheet(self, metrics: List[FinancialMetric]):
        """Add cash flow sheet."""
        ws = self.wb.create_sheet("Cash Flow")
        
        # Filter cash flow metrics
        cf_keywords = ["cash flow", "operating cash", "investing cash", "financing cash", "free cash", "capex"]
        cf_metrics = [
            m for m in metrics
            if any(kw in m.metric_name.lower() for kw in cf_keywords)
        ]
        
        if not cf_metrics:
            ws['A1'] = "No cash flow data available"
            return
        
        # Create table
        self._create_statement_table(ws, "Cash Flow Statement", cf_metrics)
    
    def _add_ratios_sheet(self, derived_metrics: List[FinancialMetric]):
        """Add ratios and analysis sheet."""
        ws = self.wb.create_sheet("Ratios & Analysis")
        
        if not derived_metrics:
            ws['A1'] = "No ratio data available"
            return
        
        # Create table
        self._create_statement_table(ws, "Financial Ratios", derived_metrics)
    
    def _add_raw_data_sheet(
        self,
        metrics: List[FinancialMetric],
        derived_metrics: Optional[List[FinancialMetric]]
    ):
        """Add raw data sheet."""
        ws = self.wb.create_sheet("Raw Data")
        
        all_metrics = metrics + (derived_metrics or [])
        
        # Headers
        headers = ["Metric ID", "Metric Name", "Value", "Currency", "Scale", "Period", "Entity Type"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
        
        # Data rows
        for row_idx, metric in enumerate(all_metrics, 2):
            ws.cell(row=row_idx, column=1, value=metric.metric_id)
            ws.cell(row=row_idx, column=2, value=metric.metric_name)
            ws.cell(row=row_idx, column=3, value=float(metric.value))
            ws.cell(row=row_idx, column=4, value=metric.currency)
            ws.cell(row=row_idx, column=5, value=metric.scale)
            ws.cell(row=row_idx, column=6, value=metric.period_end_date.strftime('%Y-%m-%d') if metric.period_end_date else "N/A")
            ws.cell(row=row_idx, column=7, value=metric.entity_type.value if metric.entity_type else "N/A")
        
        # Format columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_statement_table(
        self,
        ws,
        title: str,
        metrics: List[FinancialMetric]
    ):
        """Create a financial statement table."""
        # Title
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
        
        # Group metrics by period
        from collections import defaultdict
        by_period = defaultdict(list)
        for m in metrics:
            by_period[m.period_end_date].append(m)
        
        # Get sorted periods
        periods = sorted(by_period.keys(), reverse=True)
        
        # Headers
        row = 3
        ws.cell(row=row, column=1, value="Metric")
        self._apply_header_style(ws.cell(row=row, column=1))
        
        for col, period in enumerate(periods, 2):
            cell = ws.cell(row=row, column=col, value=period.strftime('%Y-%m-%d') if period else "N/A")
            self._apply_header_style(cell)
        
        # Get unique metric names
        metric_names = sorted(list(set(m.metric_name for m in metrics)))
        
        # Data rows
        for metric_name in metric_names:
            row += 1
            ws.cell(row=row, column=1, value=metric_name)
            
            for col, period in enumerate(periods, 2):
                # Find metric for this period
                metric = next((m for m in by_period[period] if m.metric_name == metric_name), None)
                if metric:
                    ws.cell(row=row, column=col, value=float(metric.value))
                else:
                    ws.cell(row=row, column=col, value="N/A")
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        for col in range(2, len(periods) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _apply_header_style(self, cell):
        """Apply header styling to a cell."""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
